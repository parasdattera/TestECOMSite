from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import traceback

def write_to_excel(products, filename="products.xlsx"):
    """
    Writes a list of product dictionaries to an Excel file with timestamp.
    
    Args:
        products (list): List of dictionaries. Each dict must contain 'name', 'price', 'description'.
        filename (str): The filename to save the Excel file as (default: 'products.xlsx').
    """
    try:
        if not isinstance(products, list):
            raise TypeError("products must be a list of dictionary object.")

        if not products:
            print("No product data to write.")
        
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Data"

        headers = ["Product Name", "Price", "Description", "Timestamp"]
        ws.append(headers)

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for product in products:
            if not isinstance(product, dict):
                print(f"Warning: Skipping invalid product entry: {product}")
                continue
            ws.append([
                product.get("name", ""),
                product.get("price", ""),
                product.get("description", ""),
                timestamp
            ])

        # Auto-adjust column widths
        for col_idx, _ in enumerate(headers, 1):
            column = get_column_letter(col_idx)
            max_length = max((len(str(cell.value)) for cell in ws[column]), default=0)
            ws.column_dimensions[column].width = max(15, max_length + 2)

        # Save to Excel
        wb.save(filename)
        print(f"Data successfully written to '{filename}'")

    except PermissionError:
        print(f"Permission denied: Close the file '{filename}' if it's open in Excel.")
    except FileNotFoundError:
        print(f"Invalid path: Cannot save to '{filename}'. Check if the directory exists.")
    except OSError as oe:
        print(f"OS error: {oe}")
    except TypeError as te:
        print(f"Type error: {te}")
    except Exception as e:
        print("An unexpected error occurred while writing to Excel:")
        traceback.print_exc()
